import time
import discord
import datetime
from time import sleep
from openpyxl import load_workbook, Workbook
import shutil
import os

# 自分のBotのアクセストークンに置き換え
TOKEN = ''

# 接続に必要なオブジェクトを生成
client = discord.Client(intents=discord.Intents.all())

Sunday = "<:Sunday:1234567>"
Monday = ""
Tuesday = ""
Wednesday = ""
Thursday = ""
Friday = ""
Saturday = ""
date = [Sunday, Monday, Tuesday, Wednesday, Thursday, Friday, Saturday]

ID_CHANNEL_SHIFT =
ID_CHANNEL_only_managers =

# スクリプトのディレクトリを基準に相対パスを設定
base_dir = os.path.dirname(os.path.abspath(__file__))

# Discordはニックネームで登録するので、あらかじめ本名を登録しておくテキストファイルを読み込む
with open(os.path.join(base_dir, 'name.txt'), 'r', encoding='utf-8') as f:
    name_list = f.readlines()
    name_list = [name.rstrip("\n") for name in name_list]
    name_dict = {}
    for name in name_list:
        name = name.split(" ")
        name_dict[name[0]] = name[1]
print(name_dict)

# 次の週が何月何日から何月何日までかを現時刻から計算
now = datetime.datetime.now()
if now.weekday() == 6:
    next_sunday = now + datetime.timedelta(days=7)
else:
    next_sunday = now + datetime.timedelta(days=(6-now.weekday()))
next_saturday = next_sunday + datetime.timedelta(days=6)
next_sunday_str = next_sunday.strftime("%m月%d日")
next_saturday_str = next_saturday.strftime("%m月%d日")

# 募集の状態を管理するファイル
STATUS_FILE = os.path.join(base_dir, 'recruitment_status.txt')
LAST_RESET_FILE = os.path.join(base_dir, 'last_reset.txt')

# 状態を読み込む関数
def read_status():
    try:
        with open(STATUS_FILE, 'r', encoding='utf-8') as f:
            status = f.readline().strip().split(": ")[1]
    except FileNotFoundError:
        status = "募集前"
        write_status(status)
    return status

# 状態を書き込む関数
def write_status(status):
    with open(STATUS_FILE, 'w', encoding='utf-8') as f:
        f.write(f"status: {status}")

# 最終リセット日を読み込む関数
def read_last_reset():
    try:
        with open(LAST_RESET_FILE, 'r', encoding='utf-8') as f:
            last_reset = f.readline().strip()
    except FileNotFoundError:
        last_reset = ""
    return last_reset

# 最終リセット日を書き込む関数
def write_last_reset(date_str):
    with open(LAST_RESET_FILE, 'w', encoding='utf-8') as f:
        f.write(date_str)

# 起動時に動作する処理
@client.event
async def on_ready():
    channel = client.get_channel(ID_CHANNEL_only_managers)
    current_status = read_status()

    # 最終リセット日を確認し、日曜日が更新された場合にリセット
    last_reset = read_last_reset()
    today_str = now.strftime("%Y-%m-%d")
    if now.weekday() == 6 and today_str != last_reset:
        write_status("募集前")
        write_last_reset(today_str)
        current_status = "募集前"

    await channel.send(f"Botが起動しました。\n現在の募集状況：{current_status}\nこのチャンネルで「シフト」というメッセージを送信すると、次の週のシフト希望を募集するかどうかを確認します。\nまた、「締め切り」というメッセージを送信すると、シフト希望の募集を締め切ります。")

@client.event
async def on_message(message):
    global is_closed, is_recruiting
    if message.channel.id == ID_CHANNEL_only_managers:
        if message.content == "シフト":
            current_status = read_status()
            if current_status == "締め切り済み":
                await message.channel.send("シフト希望は既に締め切られています。再募集しますか？「はい」または「いいえ」で答えてください。")
                answer = await client.wait_for("message")
                if answer.content == "はい":
                    write_status("募集中")
                    await message.channel.send("シフト希望の再募集を開始します。")
                    await restart_shift_schedule()
                else:
                    await message.channel.send("再募集は行われませんでした。")
            else:
                await next_week_shift()
        elif message.content == "締め切り":
            current_status = read_status()
            if current_status == "募集中":
                is_closed = True
                is_recruiting = False
                write_status("締め切り済み")
                await message.channel.send("シフト希望の募集を締め切りました。「シフト」を入力すると再募集できます。")
                channel_shift = client.get_channel(ID_CHANNEL_SHIFT)
                await channel_shift.send("シフト希望の募集を締め切りました。")
            else:
                await message.channel.send("現在、シフト希望の募集は行われていません。")

async def next_week_shift():
    global is_closed, is_recruiting
    is_closed = False
    channel = client.get_channel(ID_CHANNEL_only_managers)

    # すでに来週のシフト希望が募集されているかどうかを確認
    workbook = load_workbook(filename=os.path.join(base_dir, 'shift_schedule.xlsx'))
    if next_sunday_str + "~" + next_saturday_str + "シフト" in workbook.sheetnames:
        await channel.send(f"{next_sunday_str}~{next_saturday_str}のシフト希望はすでに募集されています！メッセージを確認してください！")
        is_recruiting = False
        write_status("募集中")
    else:
        is_recruiting = True
        write_status("募集中")
        await channel.send(f"{next_sunday_str}~{next_saturday_str}のシフト希望を募集しますか？ハイなら任意の文字列、中断なら「n」を入力。\n botが起動した時点で来週の募集メッセージがある場合は、リアクションを押せば動きます！")
        answer = await client.wait_for("message")
        if answer.content == "n":
            await channel.send("シフト希望の募集をキャンセルしました")
            is_recruiting = False
            write_status("募集前")
            return
        await channel.send("シフト希望の募集を開始します")
        await create_shift_schedule()

async def restart_shift_schedule():
    channel_shift = client.get_channel(ID_CHANNEL_SHIFT)
    await channel_shift.send("来週のシフト再の募集を開始します。上記メッセージにリアクションしてください。")

async def create_shift_schedule():
    workbook = load_workbook(filename=os.path.join(base_dir, 'shift_schedule.xlsx'))
    template = workbook['template']
    new_sheet = workbook.copy_worksheet(template)
    valid_next_sunday_str = next_sunday_str.replace("/", "-")
    valid_next_saturday_str = next_saturday_str.replace("/", "-")
    new_sheet.title = valid_next_sunday_str + "~" + valid_next_saturday_str + "シフト"

    next_monday = (next_sunday + datetime.timedelta(days=1)).strftime("%m月%d日")
    next_tuesday = (next_sunday + datetime.timedelta(days=2)).strftime("%m月%d日")
    next_wednesday = (next_sunday + datetime.timedelta(days=3)).strftime("%m月%d日")
    next_thursday = (next_sunday + datetime.timedelta(days=4)).strftime("%m月%d日")
    next_friday = (next_sunday + datetime.timedelta(days=5)).strftime("%m月%d日")
    new_sheet.cell(row=2, column=2, value=next_sunday_str)
    new_sheet.cell(row=2, column=4, value=next_monday)
    new_sheet.cell(row=2, column=6, value=next_tuesday)
    new_sheet.cell(row=2, column=8, value=next_wednesday)
    new_sheet.cell(row=2, column=10, value=next_thursday)
    new_sheet.cell(row=2, column=12, value=next_friday)
    new_sheet.cell(row=2, column=14, value=next_saturday_str)
    workbook.save(filename=os.path.join(base_dir, 'shift_schedule.xlsx'))

    channel_shift = client.get_channel(ID_CHANNEL_SHIFT)
    lunch = await channel_shift.send(f"{next_sunday_str}~{next_saturday_str}のランチのシフト希望を募集します")
    dinner = await channel_shift.send(f"{next_sunday_str}~{next_saturday_str}のディナーのシフト希望を募集します")

    lunch_id = lunch.id
    dinner_id = dinner.id
    with open(os.path.join(base_dir, 'message_id.txt'), 'w') as f:
        f.write(str(lunch_id) + "\n")
        f.write(str(dinner_id))
    sleep(3)
    for i in range(7):
        await lunch.add_reaction(date[i])
        await dinner.add_reaction(date[i])

async def update_check_file():
    # エクセルファイルを"shift_確認用.xlsx"にコピー
    shutil.copyfile(os.path.join(base_dir, 'shift_schedule.xlsx'), os.path.join(base_dir, '../shift_確認用.xlsx'))

@client.event
async def on_raw_reaction_add(payload):
    current_status = read_status()
    if current_status != "募集中":
        return
    with open(os.path.join(base_dir, 'message_id.txt'), 'r', encoding='utf-8') as f:
        lunch_id = int(f.readline())
        dinner_id = int(f.readline())
    if payload.message_id == lunch_id:
        member = client.get_user(payload.user_id)
        if member.bot:
            return
        member_name = member.name
        await record_shift(payload, member_name, "ランチ")
    elif payload.message_id == dinner_id:
        member = client.get_user(payload.user_id)
        if member.bot:
            return
        member_name = member.name
        await record_shift(payload, member_name, "ディナー")

@client.event
async def on_raw_reaction_remove(payload):
    current_status = read_status()
    if current_status != "募集中":
        return
    with open(os.path.join(base_dir, 'message_id.txt'), 'r', encoding='utf-8') as f:
        lunch_id = int(f.readline())
        dinner_id = int(f.readline())
    if payload.message_id == lunch_id:
        member = client.get_user(payload.user_id)
        if member.bot:
            return
        member_name = member.name
        await delete_shift(payload, member_name, "ランチ")
    elif payload.message_id == dinner_id:
        member = client.get_user(payload.user_id)
        if member.bot:
            return
        member_name = member.name
        await delete_shift(payload, member_name, "ディナー")

async def record_shift(payload, member_name, shift):
    member_name = name_dict[member_name]
    print(member_name + "さんが" + payload.emoji.name + "の" + shift + "のシフト希望を出しました")
    workbook = load_workbook(filename=os.path.join(base_dir, 'shift_schedule.xlsx'))
    worksheet = workbook[next_sunday_str + "~" + next_saturday_str + "シフト"]
    if payload.emoji.name == "Sunday":
        day = 2
    elif payload.emoji.name == "Monday":
        day = 4
    elif payload.emoji.name == "Tuesday":
        day = 6
    elif payload.emoji.name == "Wednesday":
        day = 8
    elif payload.emoji.name == "Thursday":
        day = 10
    elif payload.emoji.name == "Friday":
        day = 12
    elif payload.emoji.name == "Saturday":
        day = 14
    if shift == "ランチ":
        for i in range(3, 10):
            if worksheet.cell(row=i, column=day).value is None:
                worksheet.cell(row=i, column=day, value=member_name)
                break
    elif shift == "ディナー":
        for i in range(10, 17):
            if worksheet.cell(row=i, column=day).value is None:
                worksheet.cell(row=i, column=day, value=member_name)
                break
    workbook.save(filename=os.path.join(base_dir, 'shift_schedule.xlsx'))
    await update_check_file()

async def delete_shift(payload, member_name, shift):
    member_name = name_dict[member_name]
    print(member_name + "さんが" + payload.emoji.name + "の" + shift + "のシフト希望を取り消しました")
    workbook = load_workbook(filename=os.path.join(base_dir, 'shift_schedule.xlsx'))
    worksheet = workbook[next_sunday_str + "~" + next_saturday_str + "シフト"]
    if payload.emoji.name == "Sunday":
        day = 2
    elif payload.emoji.name == "Monday":
        day = 4
    elif payload.emoji.name == "Tuesday":
        day = 6
    elif payload.emoji.name == "Wednesday":
        day = 8
    elif payload.emoji.name == "Thursday":
        day = 10
    elif payload.emoji.name == "Friday":
        day = 12
    elif payload.emoji.name == "Saturday":
        day = 14
    if shift == "ランチ":
        for i in range(3, 10):
            if worksheet.cell(row=i, column=day).value == member_name:
                worksheet.cell(row=i, column=day, value="")
                break
    elif shift == "ディナー":
        for i in range(10, 17):
            if worksheet.cell(row=i, column=day).value == member_name:
                worksheet.cell(row=i, column=day, value="")
                break
    workbook.save(filename=os.path.join(base_dir, 'shift_schedule.xlsx'))
    await update_check_file()

# Botの起動とDiscordサーバーへの接続
client.run(TOKEN)
